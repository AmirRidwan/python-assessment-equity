from pathlib import Path

from openpyxl import load_workbook


def test_generated_report_structure():
    report_dir = Path("reports")

    files = list(report_dir.glob("*.xlsx"))

    assert files, "No generated Excel report was found"

    latest_report = max(
        files,
        key=lambda path: path.stat().st_mtime,
    )

    workbook = load_workbook(latest_report)

    assert "Portfolio Report" in workbook.sheetnames

    worksheet = workbook["Portfolio Report"]

    headers = [cell.value for cell in worksheet[1]]

    expected_headers = [
        "Date",
        "Ticker",
        "Price",
        "5-Day MA",
        "20-Day MA",
        "% Change",
        "Target Weight %",
        "Signal",
        "Note",
    ]

    assert headers == expected_headers

    assert worksheet.max_row >= 1

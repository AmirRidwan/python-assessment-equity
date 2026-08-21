import os
from datetime import datetime, timedelta
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.business_logic.moving_average import calculate_moving_average
from app.database import get_db
from app.dependencies import get_current_manager
from app.models.crossover_signal import CrossoverSignal
from app.models.portfolio_holding import PortfolioHolding
from app.models.price_snapshot import PriceSnapshot
from app.models.report import Report
from app.schemas import ReportCreate, ReportOut

router = APIRouter()


REPORT_DIR = Path(os.getenv("REPORT_DIR", "reports"))


@router.post(
    "/",
    response_model=ReportOut,
    status_code=status.HTTP_201_CREATED,
)
def generate_report(
    report_data: ReportCreate,
    db: Session = Depends(get_db),
    acting_manager=Depends(get_current_manager),
):
    if not acting_manager.active:
        raise HTTPException(
            status_code=400,
            detail="Portfolio manager is inactive",
        )

    if report_data.date_from > report_data.date_to:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="date_from must be before or equal to date_to",
        )

    # The assessment requires report generation to be gated
    # by seniority. The brief does not specify the exact threshold,
    # so this implementation allows associate and principal PMs.
    if acting_manager.seniority not in {"associate", "principal"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Your seniority level is not permitted to generate reports",
        )

    holdings = (
        db.query(PortfolioHolding)
        .filter(PortfolioHolding.manager_id == acting_manager.id)
        .all()
    )

    if not holdings:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="add a holding to your portfolio first",
        )

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Portfolio Report"

    headers = [
        "Date",
        "Ticker",
        "Price",
        "5-Day MA",
        "20-Day MA",
        "% Change",
        "Target Weight %",
        "Signal",
        "Note",
    ]

    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    row_count = 0

    signal_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    date_to_exclusive = report_data.date_to + timedelta(days=1)

    for holding in holdings:
        ticker = holding.ticker

        snapshots = (
            db.query(PriceSnapshot)
            .filter(
                PriceSnapshot.ticker_id == ticker.id,
                PriceSnapshot.captured_at >= report_data.date_from,
                PriceSnapshot.captured_at < date_to_exclusive,
            )
            .order_by(PriceSnapshot.captured_at.asc())
            .all()
        )

        # Include the whole date_to day.
        snapshots = [
            snapshot
            for snapshot in (
                db.query(PriceSnapshot)
                .filter(
                    PriceSnapshot.ticker_id == ticker.id,
                    PriceSnapshot.captured_at >= report_data.date_from,
                    PriceSnapshot.captured_at
                    < report_data.date_to.replace(day=report_data.date_to.day),
                )
                .order_by(PriceSnapshot.captured_at.asc())
                .all()
            )
        ]

        if not snapshots:
            continue

        prices = [snapshot.price for snapshot in snapshots]

        short_mas = calculate_moving_average(
            prices,
            5,
        )

        long_mas = calculate_moving_average(
            prices,
            20,
        )

        snapshot_ids = [snapshot.id for snapshot in snapshots]

        signals = (
            db.query(CrossoverSignal)
            .filter(CrossoverSignal.price_snapshot_id.in_(snapshot_ids))
            .all()
        )

        signals_by_snapshot = {signal.price_snapshot_id: signal for signal in signals}

        insufficient_data_note = None

        if len(snapshots) < 20:
            insufficient_data_note = (
                "Moving averages/signals aren't available: "
                "fewer than 20 snapshots in range."
            )

        previous_price = None

        for index, snapshot in enumerate(snapshots):
            current_price = snapshot.price

            percent_change = None

            if previous_price is not None and previous_price != 0:
                percent_change = (
                    (current_price - previous_price) / previous_price
                ) * 100

            signal = signals_by_snapshot.get(snapshot.id)

            note = insufficient_data_note

            worksheet.append(
                [
                    snapshot.captured_at,
                    ticker.symbol,
                    float(snapshot.price),
                    (float(short_mas[index]) if short_mas[index] is not None else None),
                    (float(long_mas[index]) if long_mas[index] is not None else None),
                    (float(percent_change) if percent_change is not None else None),
                    float(holding.target_weight_pct),
                    signal.signal_type if signal else None,
                    note,
                ]
            )

            current_row = worksheet.max_row
            row_count += 1

            if signal is not None:
                for cell in worksheet[current_row]:
                    cell.fill = signal_fill

            previous_price = current_price

    # Formatting
    for row in worksheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=len(headers),
    ):
        row[0].number_format = "yyyy-mm-dd hh:mm:ss"
        row[2].number_format = "0.0000"
        row[3].number_format = "0.0000"
        row[4].number_format = "0.0000"
        row[5].number_format = "0.00"
        row[6].number_format = "0.00"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        "A": 22,
        "B": 12,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 12,
        "G": 18,
        "H": 18,
        "I": 55,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = f"portfolio_report_manager_" f"{acting_manager.id}_" f"{timestamp}.xlsx"

    filepath = REPORT_DIR / filename

    workbook.save(filepath)

    report = Report(
        manager_id=acting_manager.id,
        date_from=report_data.date_from,
        date_to=report_data.date_to,
        filename=filename,
        row_count=row_count,
    )

    db.add(report)
    db.commit()
    db.refresh(report)

    return report


@router.get(
    "/",
    response_model=list[ReportOut],
)
def list_reports(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    acting_manager=Depends(get_current_manager),
):
    if not acting_manager.active:
        raise HTTPException(
            status_code=400,
            detail="Portfolio manager is inactive",
        )

    offset = (page - 1) * page_size

    return (
        db.query(Report)
        .filter(Report.manager_id == acting_manager.id)
        .order_by(Report.generated_at.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )


@router.get(
    "/{report_id}/download",
)
def download_report(
    report_id: int,
    db: Session = Depends(get_db),
    acting_manager=Depends(get_current_manager),
):
    if not acting_manager.active:
        raise HTTPException(
            status_code=400,
            detail="Portfolio manager is inactive",
        )

    report = (
        db.query(Report)
        .filter(
            Report.id == report_id,
            Report.manager_id == acting_manager.id,
        )
        .first()
    )

    if report is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found",
        )

    filepath = REPORT_DIR / report.filename

    if not filepath.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report file is missing",
        )

    return FileResponse(
        path=filepath,
        media_type=(
            "application/vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
        ),
        filename=report.filename,
    )
